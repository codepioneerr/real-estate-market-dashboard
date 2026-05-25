"""
Real Estate Market Dashboard — Excel workbook builder.
All content spans 12 columns (A:L) consistently: headers, tables, and charts.
Charts embedded as matplotlib PNGs — renders in Numbers, Excel, Google Sheets.
"""

import io, os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE      = os.path.join(os.path.dirname(__file__), "..")
CLEAN_DIR = os.path.join(BASE, "data", "cleaned")
DASH_DIR  = os.path.join(BASE, "dashboard")
os.makedirs(DASH_DIR, exist_ok=True)
OUTPUT    = os.path.join(DASH_DIR, "Real_Estate_Market_Dashboard.xlsx")

# ── Palette ───────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E5299"
LIGHT_BLUE = "BDD7EE"
GOLD       = "C9A227"
WHITE      = "FFFFFF"
LT_GRAY    = "F2F2F2"
DARK_GRAY  = "404040"
GRN_DARK   = "375623"
GRN_FILL   = "E2EFDA"
RED_FILL   = "FCE4D6"
RED_TEXT   = "C00000"
OFF_WHITE  = "F8F9FA"

M_BLUE  = "#2E5299"
M_GOLD  = "#C9A227"
M_GREEN = "#375623"

CITIES       = ["New York City", "Washington DC", "Maryland Suburbs"]
CITIES_SHORT = ["NYC", "Washington DC", "MD Suburbs"]
CITY_XL      = [MID_BLUE, GOLD, GRN_DARK]
CITY_MP      = [M_BLUE, M_GOLD, M_GREEN]

# Layout — 12 columns, each 15 chars wide ≈ 1260px total
NCOLS   = 12
COL_W   = 15
# Table column spans: 4 logical cols × 3 sheet cols = 12
TS = [3, 3, 3, 3]

# Chart pixel widths calibrated to 12 × 15-char cols ≈ 1260px at 96 DPI
CW_FULL  = 1230   # full-width chart
CW_HALF  = 600    # half-width (side-by-side pair)
CH_STD   = 290    # standard height
CH_TALL  = 315    # tall (trend chart)
CH_SHORT = 210    # short (affordability)


# ── Excel helpers ─────────────────────────────────────────────────────────────

def C(n):   return get_column_letter(n)
def xf(h):  return PatternFill("solid", fgColor=h)
def bf(sz=11, col=WHITE): return Font(bold=True, size=sz, color=col, name="Calibri")
def rf(sz=10, col=DARK_GRAY): return Font(size=sz, color=col, name="Calibri")
def ca(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def la(wrap=False): return Alignment(horizontal="left",   vertical="center", wrap_text=wrap, indent=1)

def thin():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def row_h(ws, r, h): ws.row_dimensions[r].height = h
def spacer(ws, r, h=8): row_h(ws, r, h)

def set_widths(ws):
    for i in range(1, NCOLS + 1):
        ws.column_dimensions[C(i)].width = COL_W

def _hmerge(ws, row, c1, c2):
    if c2 > c1:
        ws.merge_cells(f"{C(c1)}{row}:{C(c2)}{row}")
    return ws.cell(row=row, column=c1)

# ── Banner / header cells ─────────────────────────────────────────────────────

def banner(ws, row, text, bg=DARK_BLUE, fg=WHITE, sz=15, h=38):
    cell = _hmerge(ws, row, 1, NCOLS)
    cell.value = text
    cell.font  = Font(bold=True, size=sz, color=fg, name="Calibri")
    cell.fill  = xf(bg)
    cell.alignment = ca()
    row_h(ws, row, h)

def subbanner(ws, row, text):
    cell = _hmerge(ws, row, 1, NCOLS)
    cell.value = text
    cell.font  = Font(size=10, color=WHITE, name="Calibri")
    cell.fill  = xf(MID_BLUE)
    cell.alignment = ca()
    row_h(ws, row, 18)

def sect(ws, row, text, bg=MID_BLUE, c1=1, c2=None, sz=11):
    if c2 is None: c2 = NCOLS
    cell = _hmerge(ws, row, c1, c2)
    cell.value = text.upper()
    cell.font  = Font(bold=True, size=sz, color=WHITE, name="Calibri")
    cell.fill  = xf(bg)
    cell.alignment = ca()
    row_h(ws, row, 22)

# ── KPI cards ─────────────────────────────────────────────────────────────────

def kpi(ws, r, c1, c2, label, val, bg, fg=DARK_BLUE):
    lbl = _hmerge(ws, r,   c1, c2)
    lbl.value = label; lbl.font = Font(bold=True, size=9, color=fg, name="Calibri")
    lbl.fill  = xf(bg); lbl.alignment = ca()
    row_h(ws, r, 18)
    vc = _hmerge(ws, r+1, c1, c2)
    vc.value = val; vc.font = Font(bold=True, size=19, color=fg, name="Calibri")
    vc.fill  = xf(bg); vc.alignment = ca()
    row_h(ws, r+1, 28)

# ── Full-width tables (4 logical cols × 3 sheet cols each = 12) ──────────────

def wide_hdr(ws, row, labels, colors, spans=None):
    """Header row that spans all 12 sheet columns."""
    if spans is None: spans = TS
    col = 1
    for lbl, bg, span in zip(labels, colors, spans):
        cell = _hmerge(ws, row, col, col + span - 1)
        cell.value = lbl
        cell.font  = bf(10, WHITE)
        cell.fill  = xf(bg)
        cell.alignment = ca(wrap=True)
        cell.border = thin()
        col += span
    row_h(ws, row, 22)

def wide_row(ws, row, values, bgs, fcs=None, aligns=None, bolds=None, spans=None, h=20):
    """Data row that spans all 12 sheet columns."""
    if spans  is None: spans  = TS
    if fcs    is None: fcs    = [DARK_GRAY] * len(values)
    if aligns is None: aligns = [ca()]      * len(values)
    if bolds  is None: bolds  = [False]     * len(values)
    col = 1
    for val, bg, fc, align, bold, span in zip(values, bgs, fcs, aligns, bolds, spans):
        cell = _hmerge(ws, row, col, col + span - 1)
        cell.value     = val
        cell.font      = Font(bold=bold, size=10, color=fc, name="Calibri")
        cell.fill      = xf(bg)
        cell.alignment = align
        cell.border    = thin()
        col += span
    row_h(ws, row, h)


# ── Matplotlib helpers ────────────────────────────────────────────────────────

def _style(ax, ylabel=None, xlabel=None):
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#DEDEDE")
    ax.spines["bottom"].set_color("#DEDEDE")
    ax.tick_params(colors="#555", labelsize=9, length=0)
    ax.set_facecolor("white")
    ax.set_axisbelow(True)
    ax.yaxis.grid(True, color="#F0F0F0", linewidth=0.8)
    if ylabel: ax.set_ylabel(ylabel, fontsize=9, color="#555", labelpad=5)
    if xlabel: ax.set_xlabel(xlabel, fontsize=9, color="#555", labelpad=5)

def _img(fig, w, h):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    buf.seek(0); plt.close(fig)
    img = XLImage(buf)
    img.width, img.height = w, h
    return img

def _legend(ax, loc="upper left"):
    from matplotlib.lines import Line2D
    handles = [Line2D([0],[0], color=c, lw=3, label=l)
               for c, l in zip(CITY_MP, CITIES)]
    ax.legend(handles=handles, fontsize=9, loc=loc,
              framealpha=0.9, edgecolor="#DDD", fancybox=False)


# ── Chart functions ───────────────────────────────────────────────────────────

def ch_rent_bar(rent_df):
    vals = [rent_df[rent_df["City"]==c]["AvgRent"].iloc[-1] for c in CITIES]
    fig, ax = plt.subplots(figsize=(6, 3.4))
    bars = ax.bar(CITIES_SHORT, vals, color=CITY_MP, width=0.52, zorder=3,
                  edgecolor="white", linewidth=0.4)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"${x:,.0f}"))
    ax.set_ylim(0, max(vals) * 1.22)
    _style(ax, ylabel="Avg Monthly Rent ($)")
    for bar, v in zip(bars, vals):
        ax.text(bar.get_x()+bar.get_width()/2, v+max(vals)*0.018,
                f"${v:,.0f}", ha="center", va="bottom",
                fontsize=9.5, fontweight="bold", color="#333")
    fig.tight_layout(pad=0.7)
    return _img(fig, CW_HALF, CH_STD)

def ch_vacancy_bar(vacancy_df):
    latest = vacancy_df[vacancy_df["Year"]==vacancy_df["Year"].max()]
    vals = [latest[latest["City"]==c]["VacancyRate"].values[0]
            if len(latest[latest["City"]==c]) else 0 for c in CITIES]
    fig, ax = plt.subplots(figsize=(6, 3.4))
    bars = ax.bar(CITIES_SHORT, vals, color=CITY_MP, width=0.52, zorder=3,
                  edgecolor="white", linewidth=0.4)
    ax.axhline(5, color="#CC2222", ls="--", lw=1.2, alpha=0.7,
               label="5% market threshold", zorder=4)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"{x:.1f}%"))
    ax.set_ylim(0, max(vals) * 1.38)
    _style(ax, ylabel="Vacancy Rate (%)")
    ax.legend(fontsize=8.5, framealpha=0.9, edgecolor="#DDD", fancybox=False)
    for bar, v in zip(bars, vals):
        ax.text(bar.get_x()+bar.get_width()/2, v+0.07,
                f"{v:.1f}%", ha="center", va="bottom",
                fontsize=9.5, fontweight="bold", color="#333")
    fig.tight_layout(pad=0.7)
    return _img(fig, CW_HALF, CH_STD)

def ch_rent_trend(rent_df):
    fig, ax = plt.subplots(figsize=(13, 3.8))
    for city, color in zip(CITIES, CITY_MP):
        d = rent_df[rent_df["City"]==city].sort_values("date")
        ax.plot(d["date"], d["AvgRent"], color=color, lw=2.4, label=city, zorder=3)
    ax.axvspan(pd.Timestamp("2020-03"), pd.Timestamp("2021-06"),
               alpha=0.06, color="#CC2222", label="COVID period")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"${x:,.0f}"))
    _style(ax, ylabel="Avg Monthly Rent ($)")
    _legend(ax)
    fig.tight_layout(pad=0.8)
    return _img(fig, CW_FULL, CH_TALL)

def ch_mom_line(rent_df):
    cutoff = rent_df["date"].max() - pd.DateOffset(months=23)
    recent = rent_df[rent_df["date"] >= cutoff]
    fig, ax = plt.subplots(figsize=(13, 3.4))
    for city, color in zip(CITIES, CITY_MP):
        d = recent[recent["City"]==city].sort_values("date")
        ax.plot(d["date"], d["MoMChange"], color=color, lw=2,
                marker="o", markersize=3.5, label=city, zorder=3)
    ax.axhline(0, color="#BBB", lw=0.9)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"{x:+.1f}%"))
    _style(ax, ylabel="MoM Change (%)")
    _legend(ax, "upper right")
    fig.tight_layout(pad=0.8)
    return _img(fig, CW_FULL, CH_STD)

def ch_yoy_bar(rent_df):
    df = rent_df.copy(); df["Year"] = df["date"].dt.year
    yoy   = df[df["Year"]>=2020].groupby(["City","Year"])["YoYChange"].mean().reset_index()
    years = sorted(yoy["Year"].unique())
    x = np.arange(len(years)); w = 0.25
    fig, ax = plt.subplots(figsize=(13, 3.4))
    for i, (city, color) in enumerate(zip(CITIES, CITY_MP)):
        vals = [yoy[(yoy["City"]==city)&(yoy["Year"]==y)]["YoYChange"].values[0]
                if len(yoy[(yoy["City"]==city)&(yoy["Year"]==y)]) else 0 for y in years]
        ax.bar(x+(i-1)*w, vals, w, color=color, label=city, zorder=3,
               edgecolor="white", lw=0.4)
    ax.axhline(0, color="#BBB", lw=0.9)
    ax.set_xticks(x); ax.set_xticklabels([str(y) for y in years], fontsize=9.5)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"{x:+.1f}%"))
    _style(ax, ylabel="YoY Avg Change (%)")
    _legend(ax, "upper right")
    fig.tight_layout(pad=0.8)
    return _img(fig, CW_FULL, CH_STD)

def ch_affordability(afford_df):
    vals = [afford_df[afford_df["City"]==c]["AffordabilityIndex"].values[0]
            if len(afford_df[afford_df["City"]==c]) else 0 for c in CITIES]
    fig, ax = plt.subplots(figsize=(13, 2.6))
    bars = ax.barh(CITIES_SHORT[::-1], vals[::-1], color=CITY_MP[::-1],
                   height=0.42, zorder=3, edgecolor="white")
    ax.axvline(30, color="#CC2222", ls="--", lw=1.2, alpha=0.75,
               label="30% affordability threshold", zorder=4)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"{x:.0f}%"))
    _style(ax, xlabel="Annual Rent as % of Median Household Income")
    ax.yaxis.grid(False); ax.xaxis.grid(True, color="#F0F0F0", lw=0.8)
    ax.set_xlim(0, max(vals) * 1.18)
    for bar, v in zip(bars, vals[::-1]):
        ax.text(v+0.5, bar.get_y()+bar.get_height()/2,
                f"{v:.1f}%", va="center", fontsize=9.5, fontweight="bold", color="#333")
    ax.legend(fontsize=8.5, loc="lower right",
              framealpha=0.9, edgecolor="#DDD", fancybox=False)
    fig.tight_layout(pad=0.8)
    return _img(fig, CW_FULL, CH_SHORT)

def ch_vacancy_trend(vacancy_df):
    fig, ax = plt.subplots(figsize=(13, 3.2))
    for city, color in zip(CITIES, CITY_MP):
        d = vacancy_df[vacancy_df["City"]==city].sort_values("Year")
        ax.plot(d["Year"], d["VacancyRate"], color=color, lw=2.4,
                marker="o", markersize=6, label=city, zorder=3)
        for _, r in d.iterrows():
            ax.text(r["Year"], r["VacancyRate"]+0.12, f"{r['VacancyRate']:.1f}%",
                    ha="center", va="bottom", fontsize=8,
                    color=color, fontweight="bold")
    ax.axhline(5, color="#CC2222", ls="--", lw=1.2, alpha=0.7, label="5% threshold")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"{x:.1f}%"))
    ax.set_xticks(sorted(vacancy_df["Year"].unique()))
    ax.set_xlim(vacancy_df["Year"].min()-0.3, vacancy_df["Year"].max()+0.3)
    _style(ax, ylabel="Vacancy Rate (%)")
    _legend(ax, "upper right")
    fig.tight_layout(pad=0.8)
    return _img(fig, CW_FULL, CH_STD)


# ── Sheet 1: Market Overview ──────────────────────────────────────────────────

def build_overview(wb, rent_df, vacancy_df):
    ws = wb.active
    ws.title = "1 - Market Overview"
    ws.sheet_view.showGridLines = False
    set_widths(ws)

    banner(ws, 1, "Real Estate Market Dashboard — NYC  |  Washington DC  |  Maryland Suburbs",
           bg=DARK_BLUE, sz=16, h=40)
    subbanner(ws, 2, "Market Overview  ·  Rent, Vacancy & Five-Year Trends  ·  2019 – 2024")
    spacer(ws, 3, 10)
    sect(ws, 4, "Key Metrics — Latest Month")

    rent_last = rent_df.groupby("City")["AvgRent"].last()
    yoy_last  = rent_df.groupby("City")["YoYChange"].last()
    vac_last  = vacancy_df[vacancy_df["Year"]==vacancy_df["Year"].max()]

    bgs = [LIGHT_BLUE, "FFF2CC", "E2EFDA"]
    # Rows 5-6: avg rent (3 cards × 4 cols = 12)
    for i, city in enumerate(CITIES):
        c1 = i*4+1; kpi(ws, 5, c1, c1+3, city, f"${rent_last.get(city,0):,.0f} / mo", bgs[i])
    spacer(ws, 7, 7)
    # Rows 8-9: YoY growth
    for i, city in enumerate(CITIES):
        c1 = i*4+1; kpi(ws, 8, c1, c1+3, "YoY Rent Growth", f"{yoy_last.get(city,0):+.1f}%", bgs[i])
    spacer(ws, 10, 7)
    # Rows 11-12: vacancy
    for i, city in enumerate(CITIES):
        c1  = i*4+1
        vrow = vac_last[vac_last["City"]==city]
        vac  = vrow["VacancyRate"].values[0] if len(vrow) else 0
        kpi(ws, 11, c1, c1+3, "Vacancy Rate (2024)", f"{vac:.1f}%", bgs[i])

    spacer(ws, 13, 12)
    sect(ws, 14, "Average Rent by City — Latest Month",  c1=1, c2=6)
    sect(ws, 14, "Vacancy Rate by City — 2024",          c1=7, c2=12)
    print("    generating rent bar chart...")
    ws.add_image(ch_rent_bar(rent_df), "A15")
    print("    generating vacancy bar chart...")
    ws.add_image(ch_vacancy_bar(vacancy_df), "G15")

    spacer(ws, 29, 12)
    sect(ws, 30, "Rent Trend Over Time — All Three Markets  (Monthly, Jan 2019 – Dec 2024)")
    print("    generating rent trend chart...")
    ws.add_image(ch_rent_trend(rent_df), "A31")


# ── Sheet 2: Rent Trends ──────────────────────────────────────────────────────

def build_rent_trends(wb, rent_df):
    ws = wb.create_sheet("2 - Rent Trends")
    ws.sheet_view.showGridLines = False
    set_widths(ws)

    banner(ws, 1, "Rent Trends — Month-over-Month & Year-over-Year Analysis",
           bg=DARK_BLUE, sz=14, h=36)
    subbanner(ws, 2, "Last 24 Months MoM  ·  Annual YoY Averages  ·  2020 – 2024")
    spacer(ws, 3, 10)
    sect(ws, 4, "Month-over-Month Rent Change (%) — Last 24 Months")

    recent    = rent_df[rent_df["date"] >= rent_df["date"].max()-pd.DateOffset(months=23)].copy()
    mom_pivot = recent.pivot(index="date", columns="City",
                             values="MoMChange").reset_index().sort_values("date")

    tbl = 5
    wide_hdr(ws, tbl,
             ["Date", "New York City", "Washington DC", "Maryland Suburbs"],
             [DARK_BLUE, MID_BLUE, GOLD, GRN_DARK])

    for ri, (_, row) in enumerate(mom_pivot.iterrows(), start=1):
        r = tbl + ri
        nyc = row.get("New York City")
        dc  = row.get("Washington DC")
        md  = row.get("Maryland Suburbs")
        def _v(v): return round(v, 2) if pd.notna(v) else "—"
        def _bg(v): return (GRN_FILL if v >= 0 else RED_FILL) if pd.notna(v) else WHITE
        def _fc(v): return ("375623" if v >= 0 else RED_TEXT) if pd.notna(v) else DARK_GRAY
        def _bo(v): return pd.notna(v) and abs(v) > 0.5
        wide_row(ws, r,
                 [row["date"].strftime("%b %Y"), _v(nyc), _v(dc), _v(md)],
                 bgs=[LT_GRAY, _bg(nyc), _bg(dc), _bg(md)],
                 fcs=[DARK_GRAY, _fc(nyc), _fc(dc), _fc(md)],
                 bolds=[False, _bo(nyc), _bo(dc), _bo(md)])

    mom_chart_row = tbl + len(mom_pivot) + 2
    spacer(ws, tbl+len(mom_pivot)+1, 10)
    sect(ws, mom_chart_row, "Month-over-Month Rent Change — Trend Line (Last 24 Months)")
    print("    generating MoM chart...")
    ws.add_image(ch_mom_line(rent_df), f"A{mom_chart_row+1}")

    # YoY section
    yoy_start = mom_chart_row + 22
    spacer(ws, mom_chart_row+21, 12)
    sect(ws, yoy_start, "Year-over-Year Rent Change — Annual Average by City (2020 – 2024)")

    df2 = rent_df.copy(); df2["Year"] = df2["date"].dt.year
    yoy_p = df2.groupby(["City","Year"])["YoYChange"].mean().reset_index()
    yoy_p = yoy_p.pivot(index="Year", columns="City",
                        values="YoYChange").reset_index()
    yoy_p = yoy_p[yoy_p["Year"] >= 2020]

    ytbl = yoy_start + 1
    wide_hdr(ws, ytbl,
             ["Year", "New York City", "Washington DC", "Maryland Suburbs"],
             [DARK_BLUE, MID_BLUE, GOLD, GRN_DARK])

    for ri, (_, row) in enumerate(yoy_p.iterrows(), start=1):
        r = ytbl + ri
        nyc = row.get("New York City")
        dc  = row.get("Washington DC")
        md  = row.get("Maryland Suburbs")
        def _v(v): return round(v, 1) if pd.notna(v) else "—"
        def _bg(v): return (GRN_FILL if v >= 0 else RED_FILL) if pd.notna(v) else WHITE
        def _fc(v): return ("375623" if v >= 0 else RED_TEXT) if pd.notna(v) else DARK_GRAY
        wide_row(ws, r,
                 [int(row["Year"]), _v(nyc), _v(dc), _v(md)],
                 bgs=[LT_GRAY, _bg(nyc), _bg(dc), _bg(md)],
                 fcs=[DARK_GRAY, _fc(nyc), _fc(dc), _fc(md)],
                 bolds=[False, True, True, True])

    yoy_chart_row = ytbl + len(yoy_p) + 2
    spacer(ws, ytbl+len(yoy_p)+1, 10)
    sect(ws, yoy_chart_row, "Year-over-Year Rent Change — Grouped Bar Chart (2020 – 2024)")
    print("    generating YoY chart...")
    ws.add_image(ch_yoy_bar(rent_df), f"A{yoy_chart_row+1}")


# ── Sheet 3: Market Comparison ────────────────────────────────────────────────

def build_comparison(wb, rent_df, vacancy_df, afford_df):
    ws = wb.create_sheet("3 - Market Comparison")
    ws.sheet_view.showGridLines = False
    set_widths(ws)

    banner(ws, 1, "Market Comparison — Side-by-Side City Analysis",
           bg=DARK_BLUE, sz=14, h=36)
    subbanner(ws, 2, "NYC  ·  Washington DC  ·  Maryland Suburbs  |  Key Metrics & Indicators")
    spacer(ws, 3, 10)
    sect(ws, 4, "City Comparison — Key Metrics at a Glance")

    rent_last = rent_df.groupby("City")["AvgRent"].last()
    yoy_last  = rent_df.groupby("City")["YoYChange"].last()
    mom_last  = rent_df.groupby("City")["MoMChange"].last()
    vac_last  = vacancy_df[vacancy_df["Year"]==vacancy_df["Year"].max()].set_index("City")
    aff       = afford_df.set_index("City")

    # Table header — spans all 12 cols (3 each)
    wide_hdr(ws, 5,
             ["Metric", "New York City", "Washington DC", "Maryland Suburbs"],
             [DARK_BLUE, MID_BLUE, GOLD, GRN_DARK])

    row_bgs = [OFF_WHITE, LT_GRAY]
    metrics = [
        ("Avg Monthly Rent",
            [f"${rent_last.get(c,0):,.0f}" for c in CITIES]),
        ("YoY Rent Growth",
            [f"{yoy_last.get(c,0):+.1f}%" for c in CITIES]),
        ("MoM Rent Change",
            [f"{mom_last.get(c,0):+.2f}%" for c in CITIES]),
        ("Vacancy Rate (2024)",
            [f"{vac_last.loc[c,'VacancyRate']:.1f}%" if c in vac_last.index else "—" for c in CITIES]),
        ("Rent-to-Income Ratio",
            [f"{aff.loc[c,'AffordabilityIndex']:.1f}%" if c in aff.index else "—" for c in CITIES]),
        ("Median Household Income",
            [f"${aff.loc[c,'MedianHouseholdIncome']:,.0f}" if c in aff.index else "—" for c in CITIES]),
        ("Market Classification",
            ["Major Gateway Market", "Major Metro Market", "Suburban Market"]),
        ("Demand Pressure",
            ["Very High", "High", "Moderate-High"]),
        ("Investment Profile",
            ["High-Value / Higher Risk", "Stable Core Hold", "Value-Add / Growth"]),
    ]

    for ri, (label, vals) in enumerate(metrics):
        r   = 6 + ri
        rbg = row_bgs[ri % 2]
        wide_row(ws, r,
                 [label] + vals,
                 bgs   =[LT_GRAY, rbg, rbg, rbg],
                 aligns=[la(), ca(), ca(), ca()],
                 h=22)

    spacer(ws, 6+len(metrics), 12)

    aff_row = 6 + len(metrics) + 1
    sect(ws, aff_row, "Affordability Index — Annual Rent as % of Median Household Income")
    print("    generating affordability chart...")
    ws.add_image(ch_affordability(afford_df), f"A{aff_row+1}")

    spacer(ws, aff_row+15, 12)

    vac_sect = aff_row + 16
    sect(ws, vac_sect, "Vacancy Rate Trend — All Markets  (2020 – 2024)")

    vac_p = vacancy_df.pivot(index="Year", columns="City",
                              values="VacancyRate").reset_index()
    vtbl = vac_sect + 1
    wide_hdr(ws, vtbl,
             ["Year", "New York City", "Washington DC", "Maryland Suburbs"],
             [DARK_BLUE, MID_BLUE, GOLD, GRN_DARK])

    for ri, (_, row) in enumerate(vac_p.iterrows(), start=1):
        r = vtbl + ri
        nyc = row.get("New York City")
        dc  = row.get("Washington DC")
        md  = row.get("Maryland Suburbs")
        wide_row(ws, r,
                 [int(row["Year"]),
                  round(float(nyc),1) if pd.notna(nyc) else "—",
                  round(float(dc),1)  if pd.notna(dc)  else "—",
                  round(float(md),1)  if pd.notna(md)  else "—"],
                 bgs=[LT_GRAY, WHITE, WHITE, WHITE])

    vac_chart = vtbl + len(vac_p) + 2
    spacer(ws, vtbl+len(vac_p)+1, 10)
    sect(ws, vac_chart, "Vacancy Rate Trend — Line Chart (2020 – 2024)")
    print("    generating vacancy trend chart...")
    ws.add_image(ch_vacancy_trend(vacancy_df), f"A{vac_chart+1}")


# ── Sheet 4: Executive Summary ────────────────────────────────────────────────

def build_exec_summary(wb, rent_df, vacancy_df, afford_df):
    ws = wb.create_sheet("4 - Executive Summary")
    ws.sheet_view.showGridLines = False
    set_widths(ws)

    banner(ws, 1, "Executive Summary — Multifamily Market Intelligence Report",
           bg=DARK_BLUE, sz=16, h=40)
    subbanner(ws, 2,
        "Prepared by: Nicholas Black  |  Markets: NYC, Washington DC, Maryland  |  Analysis Period: 2019 – 2024")
    spacer(ws, 3, 10)
    sect(ws, 4, "Portfolio Snapshot — Current Month")

    rent_last = rent_df.groupby("City")["AvgRent"].last()
    yoy_last  = rent_df.groupby("City")["YoYChange"].last()
    vac_last  = vacancy_df[vacancy_df["Year"]==vacancy_df["Year"].max()].set_index("City")
    aff       = afford_df.set_index("City")

    # 6 KPI cards × 2 cols = exactly 12 cols (A:L)
    kpis = [
        ("NYC Avg Rent",   f"${rent_last.get('New York City',0):,.0f}",    LIGHT_BLUE),
        ("DC Avg Rent",    f"${rent_last.get('Washington DC',0):,.0f}",    "FFF2CC"),
        ("MD Avg Rent",    f"${rent_last.get('Maryland Suburbs',0):,.0f}", "E2EFDA"),
        ("NYC YoY Growth", f"{yoy_last.get('New York City',0):+.1f}%",    LIGHT_BLUE),
        ("DC YoY Growth",  f"{yoy_last.get('Washington DC',0):+.1f}%",    "FFF2CC"),
        ("MD YoY Growth",  f"{yoy_last.get('Maryland Suburbs',0):+.1f}%", "E2EFDA"),
    ]
    for i, (lbl, val, bg) in enumerate(kpis):
        c1 = i*2+1; kpi(ws, 5, c1, c1+1, lbl, val, bg)

    spacer(ws, 7, 10)
    sect(ws, 8, "Key Market Insights")

    nyc_r  = rent_last.get("New York City", 0)
    dc_r   = rent_last.get("Washington DC", 0)
    md_r   = rent_last.get("Maryland Suburbs", 0)
    nyc_y  = yoy_last.get("New York City", 0)
    dc_y   = yoy_last.get("Washington DC", 0)
    md_y   = yoy_last.get("Maryland Suburbs", 0)
    nyc_v  = vac_last.loc["New York City","VacancyRate"]   if "New York City"   in vac_last.index else 0
    dc_v   = vac_last.loc["Washington DC","VacancyRate"]   if "Washington DC"   in vac_last.index else 0
    md_af  = aff.loc["Maryland Suburbs","AffordabilityIndex"] if "Maryland Suburbs" in aff.index else 0
    nyc_af = aff.loc["New York City","AffordabilityIndex"]    if "New York City"    in aff.index else 0

    insights = [
        ("1.  Rent Recovery Post-COVID — New York City",
         f"New York City experienced the sharpest COVID-era correction (~10% decline, 2020–2021) but has since "
         f"recovered strongly. Current average rent stands at ${nyc_r:,.0f}/mo, reflecting {nyc_y:+.1f}% "
         f"year-over-year growth — the highest growth trajectory of the three markets tracked."),
        ("2.  Washington DC — Steady Appreciation with Low Volatility",
         f"Washington DC has delivered the most consistent rent growth across the full analysis period, with "
         f"minimal COVID disruption and a current average of ${dc_r:,.0f}/mo ({dc_y:+.1f}% YoY). The market "
         f"is underpinned by stable federal government and contractor employment demand."),
        ("3.  Maryland Suburbs — Best Rent-to-Income Value in the Region",
         f"At ${md_r:,.0f}/mo, the Maryland suburbs (Rockville, Bethesda, Silver Spring) offer the strongest "
         f"relative affordability with a rent-to-income ratio of {md_af:.1f}% vs. {nyc_af:.1f}% for NYC. "
         f"Continued in-migration from DC proper supports demand, with {md_y:+.1f}% YoY rent growth."),
        ("4.  Tightening Vacancy Signals a Landlord-Favorable Environment",
         f"All three markets show sustained vacancy declines — NYC at {nyc_v:.1f}% and DC at {dc_v:.1f}% as of "
         f"2024, both well below the 5% landlord-market threshold. Supply constraints and continued demand "
         f"pressure support near-term rent appreciation across all three metros."),
    ]

    row = 9
    for title, body in insights:
        # Title bar — full width
        tc = _hmerge(ws, row, 1, NCOLS)
        tc.value = title
        tc.font  = Font(bold=True, size=10, color=WHITE, name="Calibri")
        tc.fill  = xf(MID_BLUE)
        tc.alignment = la(wrap=False)
        row_h(ws, row, 22); row += 1

        # Body — 2-row vertical+horizontal merge
        ws.merge_cells(f"A{row}:{C(NCOLS)}{row+1}")
        bc = ws.cell(row=row, column=1, value=body)
        bc.font      = Font(size=10, color=DARK_GRAY, name="Calibri")
        bc.fill      = xf(OFF_WHITE)
        bc.alignment = Alignment(horizontal="left", vertical="center",
                                  wrap_text=True, indent=1)
        row_h(ws, row, 26); row_h(ws, row+1, 26)
        row += 2; spacer(ws, row, 7); row += 1

    spacer(ws, row, 10); row += 1

    # Recommendation
    sect(ws, row, "Investment Recommendation", bg=GOLD)
    row_h(ws, row, 24); row += 1

    ws.merge_cells(f"A{row}:{C(NCOLS)}{row+2}")
    rc = ws.cell(row=row, column=1,
        value=(
            "Maryland Suburbs — Most Compelling Near-Term Opportunity:  Below-average vacancy (~3.2%), "
            "sustained YoY rent growth, and a rent-to-income ratio that maintains regional affordability. "
            "Strong value-add and workforce housing play.     "
            "Washington DC — Stable Core Hold:  Predictable cash flow and government employment anchor "
            "support a low-volatility defensive allocation.     "
            "New York City — Highest Upside / Higher Capital Threshold:  Post-COVID recovery is strong, "
            "but a 62%+ rent-to-income ratio limits the renter pool. Best suited for institutional capital "
            "with longer hold horizons."
        ))
    rc.font = Font(size=10, color="2C2C2C", name="Calibri")
    rc.fill = xf("FFFBF0")
    rc.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=2)
    for r in [row, row+1, row+2]:
        row_h(ws, r, 22)
    row += 4; spacer(ws, row, 12); row += 1

    # Footer
    fc = _hmerge(ws, row, 1, NCOLS)
    fc.value = ("Data Sources:  Zillow Research (ZORI / ZHVI)  |  HUD Fair Market Rents  "
                "|  U.S. Census Bureau ACS  |  FRED Economic Data")
    fc.font  = Font(italic=True, size=9, color="888888", name="Calibri")
    fc.fill  = xf(LT_GRAY)
    fc.alignment = ca()
    row_h(ws, row, 18)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Loading cleaned data...")
    rent_df    = pd.read_csv(os.path.join(CLEAN_DIR, "rents_clean.csv"),   parse_dates=["date"])
    vacancy_df = pd.read_csv(os.path.join(CLEAN_DIR, "vacancy_clean.csv"))
    afford_df  = pd.read_csv(os.path.join(CLEAN_DIR, "affordability_clean.csv"))

    print("Building workbook...")
    wb = Workbook()

    print("  Sheet 1: Market Overview")
    build_overview(wb, rent_df, vacancy_df)
    print("  Sheet 2: Rent Trends")
    build_rent_trends(wb, rent_df)
    print("  Sheet 3: Market Comparison")
    build_comparison(wb, rent_df, vacancy_df, afford_df)
    print("  Sheet 4: Executive Summary")
    build_exec_summary(wb, rent_df, vacancy_df, afford_df)

    wb.save(OUTPUT)
    print(f"\nSaved: {OUTPUT}")

if __name__ == "__main__":
    main()
